import csv
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from app.enrichment.merchant_cleanup import clean_merchant
from app.parsers.base import BaseTransactionParser, FailedRow, ParsedTransaction, ParseResult
from app.shared.enums import PaymentMethod, SourceType, TransactionType


DATE_COLUMNS = ["date", "transaction date", "txn date", "paid on", "created at", "time"]
DESCRIPTION_COLUMNS = ["description", "narration", "details", "note", "remarks", "merchant", "transaction"]
MERCHANT_COLUMNS = ["merchant", "recipient", "paid to", "biller", "name"]
AMOUNT_COLUMNS = ["amount", "transaction amount", "value"]
DEBIT_COLUMNS = ["debit", "withdrawal", "paid", "money out"]
CREDIT_COLUMNS = ["credit", "deposit", "received", "money in"]
TYPE_COLUMNS = ["type", "direction", "transaction type"]
REFERENCE_COLUMNS = ["reference", "reference id", "utr", "rrn", "txn id", "transaction id"]
BALANCE_COLUMNS = ["balance", "closing balance", "available balance"]
METHOD_COLUMNS = ["payment method", "method", "mode"]


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _pick(row: dict[str, str], aliases: list[str]) -> str | None:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        if alias in normalized and str(normalized[alias]).strip():
            return str(normalized[alias]).strip()
    for key, value in normalized.items():
        if any(alias in key for alias in aliases) and str(value).strip():
            return str(value).strip()
    return None


def _parse_decimal(value: str | None) -> Decimal | None:
    if not value:
        return None
    cleaned = str(value).replace(",", "").replace("₹", "").replace("INR", "").strip()
    cleaned = cleaned.strip("()")
    if not cleaned:
        return None
    try:
        return abs(Decimal(cleaned))
    except InvalidOperation:
        return None


def _parse_date(value: str | None):
    if not value:
        raise ValueError("Missing transaction date")
    cleaned = value.strip()
    for fmt in (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d %b %Y",
        "%d %B %Y",
        "%d%b,%Y",
        "%d%B,%Y",
    ):
        try:
            return datetime.strptime(cleaned[:11], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(cleaned.replace("Z", "+00:00")).date()
    except ValueError as exc:
        raise ValueError(f"Unsupported date format: {value}") from exc


def _read_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    if Path(filename).suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [
            {headers[index]: "" if cell is None else str(cell) for index, cell in enumerate(row)}
            for row in rows[1:]
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

    text = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


class GenericCsvParser(BaseTransactionParser):
    source_type = SourceType.GENERIC_CSV
    default_payment_method = PaymentMethod.UNKNOWN

    def parse(self, filename: str, content: bytes, password: str | None = None) -> ParseResult:
        result = ParseResult()
        rows = _read_rows(filename, content)
        if not rows:
            result.warnings.append("No transaction rows were found.")
            return result

        for index, row in enumerate(rows, start=2):
            try:
                parsed = self._parse_row(row)
                result.transactions.append(parsed)
            except Exception as exc:  # noqa: BLE001
                result.failed_rows.append(FailedRow(row_number=index, message=str(exc), raw=row))
        return result

    def _parse_row(self, row: dict[str, str]) -> ParsedTransaction:
        date_value = _pick(row, DATE_COLUMNS)
        description = _pick(row, DESCRIPTION_COLUMNS) or ""
        merchant = _pick(row, MERCHANT_COLUMNS) or clean_merchant(description)
        reference_id = _pick(row, REFERENCE_COLUMNS)
        balance = _parse_decimal(_pick(row, BALANCE_COLUMNS))
        payment_method = (_pick(row, METHOD_COLUMNS) or self.default_payment_method).lower().replace(" ", "_")

        debit = _parse_decimal(_pick(row, DEBIT_COLUMNS))
        credit = _parse_decimal(_pick(row, CREDIT_COLUMNS))
        amount = _parse_decimal(_pick(row, AMOUNT_COLUMNS))
        type_value = (_pick(row, TYPE_COLUMNS) or "").lower()

        if debit is not None:
            transaction_type = TransactionType.EXPENSE
            amount = debit
        elif credit is not None:
            transaction_type = TransactionType.INCOME
            amount = credit
        elif "income" in type_value or "credit" in type_value or "received" in type_value:
            transaction_type = TransactionType.INCOME
        elif "transfer" in type_value:
            transaction_type = TransactionType.TRANSFER
        else:
            transaction_type = TransactionType.EXPENSE

        if amount is None:
            raise ValueError("Missing amount")

        return ParsedTransaction(
            transaction_date=_parse_date(date_value),
            amount=amount,
            transaction_type=transaction_type,
            payment_method=payment_method if payment_method in PaymentMethod._value2member_map_ else PaymentMethod.UNKNOWN,
            source_type=self.source_type,
            raw_description=description or merchant,
            description=description or None,
            merchant_name=merchant,
            display_name=merchant,
            reference_id=reference_id,
            balance_after_transaction=balance,
        )
